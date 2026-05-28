# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

out_path = r"C:\Users\Marketing Team\Downloads\starlent-demo\Starlent_FE_BE_Project_Plan_v3_3_utf8.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "KeHoachDuAn"

headers = [
    "STT",
    "TIÊU ĐỀ CÔNG VIỆC",
    "NGƯỜI PHỤ TRÁCH CÔNG VIỆC",
    "TRẠNG THÁI",
    "NGÀY BẮT ĐẦU",
    "NGÀY ĐẾN HẠN",
    "THỰC TẾ HOÀN THÀNH",
    "GHI CHÚ",
]

widths = [10, 86, 28, 18, 14, 14, 18, 62]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

header_row = 3
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=col, value=h)
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.fill = PatternFill("solid", fgColor="D9D9D9")

ws.row_dimensions[header_row].height = 58

thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
COLORS = {"FE": "E8F4FD", "BE": "EAF7EA", "DEVOPS": "FFF4E5", "SEC": "FDEBEC", "GEN": "FFFFFF"}

plan = [
    ("1", "Khởi động dự án và phạm vi", [
        ("1.1", "Chốt mục tiêu dự án FE/BE và tiêu chí hoàn thành", "GEN"),
        ("1.1.1", "Xác định phạm vi parity: hành vi, dữ liệu, UI, responsive", "GEN"),
        ("1.1.2", "Định nghĩa out-of-scope để tránh trôi phạm vi", "GEN"),
        ("1.2", "Tạo 2 repo ngang cấp: starlent-web, starlent-be", "GEN"),
        ("1.3", "Định nghĩa role nghiệp vụ: learner/trainer/manager/editor/admin", "GEN"),
    ]),
    ("2", "Thiết lập môi trường phát triển", [
        ("2.1", "Backend: khởi tạo FastAPI + settings/env + cấu trúc app", "BE"),
        ("2.2", "Backend: Alembic migration + schema strategy platform/tenant", "BE"),
        ("2.3", "Frontend: React + Vite + Tailwind + router + API client", "FE"),
        ("2.4", "Thiết lập .gitignore, coding standard, quy tắc branch/commit", "DEVOPS"),
    ]),
    ("3", "Backend nền tảng (FastAPI + PostgreSQL + Multi-tenant)", [
        ("3.1", "Thiết kế schema platform: tenants/users/roles/permissions/memberships", "BE"),
        ("3.2", "Thiết kế schema tenant: tenant_users/courses/modules/items/content tables", "BE"),
        ("3.3", "Auth + JWT + tenant context + permission guard", "BE"),
        ("3.4", "API nền tảng: /auth, /me, /catalog, /courses", "BE"),
        ("3.5", "API role-specific: trainer/manager/editor/admin", "BE"),
        ("3.6", "Tenant management API: list/detail/status/create/update/provision", "BE"),
    ]),
    ("4", "Frontend nền tảng + auth shell", [
        ("4.1", "Login page + session/token + logout", "FE"),
        ("4.2", "Role layout + menu + route guard", "FE"),
        ("4.3", "UI states dùng chung: Loading/Error/Empty + Toast", "FE"),
        ("4.4", "Kết nối API thật và xử lý loading/error", "FE"),
    ]),
    ("5", "Admin module parity (ưu tiên trước Editor và Learner)", [
        ("5.1", "Admin dashboard", "FE"),
        ("5.2", "Admin users: list/create + login verify user mới", "BE"),
        ("5.3", "Admin courses", "FE"),
        ("5.4", "Admin tenants + status/provision controls", "BE"),
        ("5.5", "Permission checks cho admin routes/API", "BE"),
    ]),
    ("6", "Editor module parity (sau Admin, trước Learner)", [
        ("6.1", "Editor dashboard", "FE"),
        ("6.2", "Editor content catalog", "FE"),
        ("6.3", "Editor course builder CRUD", "FE"),
        ("6.3.1", "Tạo course/module/item", "BE"),
        ("6.3.2", "Sửa/Xóa/Reorder module/item", "BE"),
        ("6.4", "Permission checks cho editor routes/API", "BE"),
    ]),
    ("7", "Learner module parity", [
        ("7.1", "Onboarding 3 bước + responsive", "FE"),
        ("7.2", "Learner dashboard", "FE"),
        ("7.3", "Courses list + detail", "FE"),
        ("7.4", "Lesson players: flashcard/video/audio/quiz/sequence/roleplay/reading", "FE"),
        ("7.5", "Assignment/Survey/Live Session", "FE"),
        ("7.6", "Engagement: leaderboard/chats/certificates/notifications/profile/settings", "FE"),
        ("7.7", "Visual parity learner theo ảnh đối chiếu", "FE"),
    ]),
    ("8", "Kiểm thử và chất lượng", [
        ("8.1", "Backend smoke_e2e", "BE"),
        ("8.2", "permission_gating_e2e", "BE"),
        ("8.3", "response_shape_e2e", "BE"),
        ("8.4", "tenant_isolation_e2e", "BE"),
        ("8.5", "Frontend build + smoke:routes", "FE"),
        ("8.6", "UAT checklist theo màn hình và role", "GEN"),
    ]),
    ("9", "Docker local stack + seed one-time", [
        ("9.1", "Dockerfile backend/frontend + docker-compose", "DEVOPS"),
        ("9.2", "Entrypoint: wait postgres -> bootstrap schema -> migrate", "DEVOPS"),
        ("9.3", "Auto-seed one-time khi DB trống, skip khi đã có data", "DEVOPS"),
        ("9.4", "Document reset volume để reseed", "DEVOPS"),
    ]),
    ("10", "Deploy DigitalOcean + CI/CD", [
        ("10.1", "Thiết kế môi trường dev/uat/prod trên DigitalOcean", "DEVOPS"),
        ("10.1.1", "Chọn hạ tầng: Droplet/App Platform/Managed DB", "DEVOPS"),
        ("10.1.2", "Thiết lập network/firewall/domain/SSL", "DEVOPS"),
        ("10.2", "Chuẩn bị deploy backend", "DEVOPS"),
        ("10.2.1", "Env secrets theo môi trường", "SEC"),
        ("10.2.2", "Migration strategy khi deploy", "DEVOPS"),
        ("10.2.3", "Healthcheck + process manager + log rotation", "DEVOPS"),
        ("10.3", "Chuẩn bị deploy frontend", "DEVOPS"),
        ("10.3.1", "Build artifact + static hosting/reverse proxy", "DEVOPS"),
        ("10.3.2", "Cấu hình API base URL theo môi trường", "DEVOPS"),
        ("10.4", "GitHub Actions CI", "DEVOPS"),
        ("10.4.1", "FE: build + smoke routes", "FE"),
        ("10.4.2", "BE: compile + e2e gates", "BE"),
        ("10.5", "GitHub Actions CD", "DEVOPS"),
        ("10.5.1", "Deploy dev tự động khi merge nhánh develop", "DEVOPS"),
        ("10.5.2", "Deploy uat theo manual approval", "DEVOPS"),
        ("10.5.3", "Deploy prod theo release tag + approval", "DEVOPS"),
        ("10.6", "Post-deploy checks + rollback checklist", "DEVOPS"),
    ]),
    ("11", "Bảo mật và chống tấn công", [
        ("11.1", "Hardening Auth/JWT", "SEC"),
        ("11.1.1", "Rút ngắn TTL access token theo môi trường", "SEC"),
        ("11.1.2", "Rotate SECRET_KEY theo quy trình", "SEC"),
        ("11.1.3", "Khóa brute-force: limit login attempts", "SEC"),
        ("11.2", "API Security baseline", "SEC"),
        ("11.2.1", "Rate limiting theo IP + endpoint nhạy cảm", "SEC"),
        ("11.2.2", "Input validation chặt (pydantic schema, length, enum)", "SEC"),
        ("11.2.3", "Chuẩn hóa error response, không lộ stacktrace", "SEC"),
        ("11.2.4", "CORS chỉ allow domain hợp lệ theo env", "SEC"),
        ("11.3", "Database & secrets", "SEC"),
        ("11.3.1", "Managed secrets, không commit .env", "SEC"),
        ("11.3.2", "Least privilege DB user theo môi trường", "SEC"),
        ("11.3.3", "Bật backup/restore test định kỳ", "SEC"),
        ("11.4", "Infra security", "SEC"),
        ("11.4.1", "Firewall: chỉ mở 80/443 public, DB private", "SEC"),
        ("11.4.2", "TLS/HTTPS bắt buộc + redirect HTTP->HTTPS", "SEC"),
        ("11.4.3", "Harden SSH (key-based, disable password nếu phù hợp)", "SEC"),
        ("11.5", "Observability & incident response", "SEC"),
        ("11.5.1", "Audit log cho hành động admin/editor quan trọng", "SEC"),
        ("11.5.2", "Alert bất thường (login fail spike, 5xx spike)", "SEC"),
        ("11.5.3", "Runbook sự cố bảo mật + quy trình rollback", "SEC"),
        ("11.6", "Secure SDLC", "SEC"),
        ("11.6.1", "Dependency scanning (pip/npm) trong CI", "SEC"),
        ("11.6.2", "SAST/secret scan trước merge", "SEC"),
        ("11.6.3", "Checklist OWASP ASVS mức cơ bản trước go-live", "SEC"),
    ]),
]

row = header_row + 1
for sec_code, sec_title, items in plan:
    ws.cell(row=row, column=1, value=sec_code)
    ws.cell(row=row, column=2, value=sec_title)
    for col in range(1, 9):
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=True, size=12)
        c.fill = PatternFill("solid", fgColor="BFBFBF")
        c.alignment = Alignment(vertical="center", horizontal="left" if col == 2 else "center", wrap_text=True)
        c.border = border
    row += 1

    for code, title, kind in items:
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=title)
        for col in range(3, 9):
            ws.cell(row=row, column=col, value="")

        fill = PatternFill("solid", fgColor=COLORS.get(kind, "FFFFFF"))
        for col in range(1, 9):
            c = ws.cell(row=row, column=col)
            c.border = border
            c.fill = fill
            if col in (1, 3, 4, 5, 6, 7):
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        row += 1

for col in range(1, 9):
    ws.cell(row=header_row, column=col).border = border

ws.cell(row=1, column=2, value="Màu phân loại task:").font = Font(bold=True)
legend = [("FE", COLORS["FE"]), ("BE", COLORS["BE"]), ("DEVOPS", COLORS["DEVOPS"]), ("SEC", COLORS["SEC"])]
for idx, (name, color) in enumerate(legend, start=3):
    c = ws.cell(row=1, column=idx, value=name)
    c.fill = PatternFill("solid", fgColor=color)
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border

ws.freeze_panes = "A4"
wb.save(out_path)
print(out_path)

