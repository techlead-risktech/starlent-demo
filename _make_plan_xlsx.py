# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

out_path = r"C:\Users\Marketing Team\Downloads\starlent-demo\Starlent_FE_BE_Project_Plan_v2.xlsx"

wb = Workbook()
ws = wb.active
ws.title = "KeHoachDuAn"

headers = [
    "STT",
    "TIÊU ĐỀ CÔNG VIỆC",
    "NGƯỜI PHỤ TRÁCH CÔNG VIỆC",
    "TRẠNG THÁI",
    "NGÀY BẮT ĐẦU",
    "NGÀY ĐẾN HẠN",
    "THỰC TẾ HOÀN THÀNH",
    "GHI CHÚ",
]

widths = [10, 78, 28, 18, 14, 14, 18, 56]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

header_row = 3
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=header_row, column=col, value=h)
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.fill = PatternFill("solid", fgColor="D9D9D9")

ws.row_dimensions[header_row].height = 58

thin = Side(style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

plan = [
    {
        "code": "1",
        "title": "Khởi động dự án và phạm vi",
        "items": [
            ("1.1", "Chốt mục tiêu dự án FE/BE và tiêu chí hoàn thành"),
            ("1.1.1", "Xác định phạm vi parity: hành vi, dữ liệu, UI, responsive"),
            ("1.1.2", "Định nghĩa out-of-scope để tránh trôi phạm vi"),
            ("1.2", "Thiết lập cấu trúc repo"),
            ("1.2.1", "Tạo 2 repo ngang cấp: starlent-web, starlent-be"),
            ("1.2.2", "Thiết lập nhánh chính, quy tắc branch và PR"),
            ("1.3", "Phân tích role nghiệp vụ"),
            ("1.3.1", "Learner"),
            ("1.3.2", "Trainer"),
            ("1.3.3", "Manager"),
            ("1.3.4", "Editor"),
            ("1.3.5", "Admin"),
        ],
    },
    {
        "code": "2",
        "title": "Thiết lập môi trường phát triển",
        "items": [
            ("2.1", "Backend skeleton"),
            ("2.1.1", "Khởi tạo FastAPI project structure"),
            ("2.1.2", "Cấu hình settings/env"),
            ("2.1.3", "Cấu hình logging cơ bản"),
            ("2.2", "Frontend skeleton"),
            ("2.2.1", "Khởi tạo React + Vite + Tailwind"),
            ("2.2.2", "Thiết lập routing base"),
            ("2.2.3", "Thiết lập API client + interceptor lỗi"),
            ("2.3", "Chuẩn phát triển"),
            ("2.3.1", "Tạo .gitignore FE/BE"),
            ("2.3.2", "Quy ước commit message"),
            ("2.3.3", "Quy tắc code style"),
        ],
    },
    {
        "code": "3",
        "title": "Backend nền tảng (FastAPI + PostgreSQL + Multi-tenant)",
        "items": [
            ("3.1", "Thiết kế schema platform"),
            ("3.1.1", "tenants/users/roles/permissions"),
            ("3.1.2", "tenant_memberships/membership_roles"),
            ("3.2", "Thiết kế schema tenant"),
            ("3.2.1", "tenant_users/courses/course_modules/course_items"),
            ("3.2.2", "content tables: flashcard/video/audio/quiz/roleplay/reading"),
            ("3.2.3", "engagement tables: chats/notifications/certificates/leaderboard"),
            ("3.3", "Auth & Permission"),
            ("3.3.1", "POST /auth/login + JWT"),
            ("3.3.2", "Context tenant qua x-tenant-code"),
            ("3.3.3", "Permission guard theo role"),
            ("3.4", "API cho learner/editor/admin"),
            ("3.4.1", "Learner API: home/courses/detail/content"),
            ("3.4.2", "Editor API: catalog/builder CRUD/reorder"),
            ("3.4.3", "Admin API: dashboard/users/courses/tenants"),
        ],
    },
    {
        "code": "4",
        "title": "Frontend nền tảng và auth shell",
        "items": [
            ("4.1", "Login + session"),
            ("4.1.1", "Form login, xử lý lỗi"),
            ("4.1.2", "Lưu token/session"),
            ("4.2", "Role layout"),
            ("4.2.1", "Sidebar/menu theo role"),
            ("4.2.2", "Route guard cho mỗi role"),
            ("4.3", "UI states dùng chung"),
            ("4.3.1", "LoadingBlock"),
            ("4.3.2", "ErrorBlock"),
            ("4.3.3", "EmptyBlock"),
            ("4.3.4", "ToastHost"),
        ],
    },
    {
        "code": "5",
        "title": "Onboarding learner (3 bước)",
        "items": [
            ("5.1", "Dựng route /onboarding"),
            ("5.1.1", "Step 1: welcome"),
            ("5.1.2", "Step 2: lộ trình"),
            ("5.1.3", "Step 3: bắt đầu học"),
            ("5.2", "Điều hướng sau login learner"),
            ("5.2.1", "Lần đầu vào onboarding"),
            ("5.2.2", "Hoàn tất thì vào dashboard"),
            ("5.3", "Responsive mobile"),
            ("5.3.1", "Viewport ~447x642"),
        ],
    },
    {
        "code": "6",
        "title": "Learner module parity",
        "items": [
            ("6.1", "Dashboard learner"),
            ("6.2", "Courses list + detail"),
            ("6.3", "Lesson players"),
            ("6.3.1", "Flashcard"),
            ("6.3.2", "Video"),
            ("6.3.3", "Audio"),
            ("6.3.4", "Quiz/Sequence"),
            ("6.3.5", "Roleplay"),
            ("6.3.6", "Reading"),
            ("6.3.7", "Assignment/Survey/Live session"),
            ("6.4", "Engagement pages"),
            ("6.4.1", "Leaderboard"),
            ("6.4.2", "Chats"),
            ("6.4.3", "Certificates"),
            ("6.4.4", "Notifications/Profile/Settings"),
        ],
    },
    {
        "code": "7",
        "title": "Editor/Admin parity",
        "items": [
            ("7.1", "Editor dashboard"),
            ("7.2", "Editor catalog"),
            ("7.3", "Editor course builder"),
            ("7.3.1", "Tạo course/module/item"),
            ("7.3.2", "Sửa/Xóa/Reorder"),
            ("7.4", "Admin dashboard"),
            ("7.5", "Admin users"),
            ("7.6", "Admin courses"),
            ("7.7", "Admin tenants + status/provision"),
        ],
    },
    {
        "code": "8",
        "title": "Kiểm thử và chất lượng",
        "items": [
            ("8.1", "Backend smoke e2e"),
            ("8.2", "Permission gating e2e"),
            ("8.3", "Response shape e2e"),
            ("8.4", "Tenant isolation e2e"),
            ("8.5", "Frontend build + smoke routes"),
            ("8.6", "UAT checklist theo màn hình"),
        ],
    },
    {
        "code": "9",
        "title": "Docker local stack + seed one-time",
        "items": [
            ("9.1", "Dockerfile backend/frontend"),
            ("9.2", "docker-compose postgres/backend/frontend"),
            ("9.3", "Entrypoint backend"),
            ("9.3.1", "Wait postgres"),
            ("9.3.2", "Bootstrap schema"),
            ("9.3.3", "Alembic migrate"),
            ("9.3.4", "Seed one-time khi DB trống"),
            ("9.4", "Document reset volume để reseed"),
        ],
    },
    {
        "code": "10",
        "title": "GitHub + CI/CD + bàn giao",
        "items": [
            ("10.1", "Init repo FE/BE + remote GitHub"),
            ("10.2", "GitHub Actions FE"),
            ("10.2.1", "Build"),
            ("10.2.2", "Smoke routes"),
            ("10.3", "GitHub Actions BE"),
            ("10.3.1", "Compile"),
            ("10.3.2", "E2E scripts"),
            ("10.4", "README hợp nhất (CMD + Docker + seed)"),
            ("10.5", "Checklist bàn giao vận hành local/UAT"),
        ],
    },
]

row = header_row + 1
for sec in plan:
    ws.cell(row=row, column=1, value=sec["code"])
    ws.cell(row=row, column=2, value=sec["title"])
    for col in range(1, 9):
        c = ws.cell(row=row, column=col)
        c.font = Font(bold=True, size=12)
        c.fill = PatternFill("solid", fgColor="BFBFBF")
        c.alignment = Alignment(vertical="center", horizontal="left" if col == 2 else "center", wrap_text=True)
        c.border = border
    row += 1

    for code, title in sec["items"]:
        ws.cell(row=row, column=1, value=code)
        ws.cell(row=row, column=2, value=title)
        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=4, value="")
        ws.cell(row=row, column=5, value="")
        ws.cell(row=row, column=6, value="")
        ws.cell(row=row, column=7, value="")
        ws.cell(row=row, column=8, value="")

        for col in range(1, 9):
            c = ws.cell(row=row, column=col)
            c.border = border
            if col == 1:
                c.alignment = Alignment(horizontal="center", vertical="center")
            elif col in (3, 4, 5, 6, 7):
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 24
        row += 1

for col in range(1, 9):
    ws.cell(row=header_row, column=col).border = border

ws.freeze_panes = "A4"
wb.save(out_path)
print(out_path)
